"""
WedstrijdCalculator - Portsmouth Yardstick (PY) Scoring Tool
============================================================
Volledig offline Python-tool voor het berekenen van zeilwedstrijdresultaten
op basis van het Portsmouth Yardstick handicap-systeem.

Gebruik:
    python wedstrijd_calculator.py                     # interactief menu
    python wedstrijd_calculator.py --invoer data.csv   # directe CSV-verwerking
    python wedstrijd_calculator.py --invoer data.xlsx  # directe Excel-verwerking
    python wedstrijd_calculator.py --demo              # run met ingebouwde voorbeelddata

PY-formule: gecorrigeerde_tijd = verlopen_tijd_seconden * 1000 / PY
Een lagere gecorrigeerde tijd = betere prestatie.
"""

import argparse
import sys
import os
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# CENTRALE PY-TABEL
# Voeg hier boottypes en hun Portsmouth Yardstick-waarden toe of pas ze aan.
# Bron: RYA Portsmouth Yardstick (meest recente nationale lijst).
# Waarden kunnen jaarlijks wijzigen - controleer de actuele RYA-lijst.
# ==============================================================================
BOAT_PY: dict[str, float] = {
    # Populaire eenpersoonsdinghies
    "Laser":            1100,   # ILCA 7 (standaard rig)
    "Laser Radial":     1147,   # ILCA 6 (radial rig)
    "Laser 4.7":        1208,   # ILCA 4 (lichtste rig)
    "RS Aero 5":        1136,   # RS Aero met 5m² zeil
    "RS Aero 6":        1098,   # RS Aero met 6m² zeil
    "RS Aero 7":        1065,   # RS Aero met 7m² zeil
    "RS Aero 9":        1014,   # RS Aero met 9m² zeil
    "RS Neo":           1176,   # RS Neo (instapper)
    "Optimist":         1646,   # jeugdboot
    "Topper":           1369,
    "Byte CII":         1135,
    "Finn":             1044,
    "Europe":           1141,
    # Tweepersoons dinghies
    "420":              1111,
    "470":              1084,
    "RS Feva XL":       1244,
    "Laser 2000":       1107,
    "Miracle":          1194,
    # Catamarans
    "Hobie Cat 16":      777,
    "Dart 18":           817,
    # Voeg hieronder eigen boottypes toe:
    # "MijnBoot": 1050,
}

# ==============================================================================
# DEMO-DATA - voorjaarswedstrijd 2026 (6 deelnemers, 5 reeksen)
# Tijden overgenomen uit het wedstrijdblad.
# PY-waarden worden automatisch opgezocht uit de BOAT_PY-tabel hierboven.
# naam, boottype, reeks, min, sec
# ==============================================================================
DEMO_DATA = [
    # naam,       boottype,          reeks, min, sec
    ("Brecht",    "Laser Radial",    1,   18,  23),
    ("Brecht",    "Laser Radial",    2,   25,  28),
    ("Brecht",    "Laser Radial",    3,   24,  43),
    ("Brecht",    "Laser Radial",    4,   23,  53),
    ("Brecht",    "Laser Radial",    5,   24,  59),

    ("Stef",      "RS Aero 7",       1,   17,  17),
    ("Stef",      "RS Aero 7",       2,   24,  14),
    ("Stef",      "RS Aero 7",       3,   50,   0),  # hoge tijd (mogelijke DNS/straf)
    ("Stef",      "RS Aero 7",       4,   23,  50),
    ("Stef",      "RS Aero 7",       5,   22,  55),

    ("Glen",      "RS Aero 7",       1,   19,   1),
    ("Glen",      "RS Aero 7",       2,   24,   2),
    ("Glen",      "RS Aero 7",       3,   23,  10),
    ("Glen",      "RS Aero 7",       4,   23,  21),
    ("Glen",      "RS Aero 7",       5,   23,  50),

    ("Karel",     "Laser",           1,   17,  45),
    ("Karel",     "Laser",           2,   25,  45),
    ("Karel",     "Laser",           3,   25,  42),
    ("Karel",     "Laser",           4,   23,  51),
    ("Karel",     "Laser",           5,   25,  39),

    ("Christine", "RS Neo",          1,   23,  36),
    ("Christine", "RS Neo",          2,   33,  45),
    ("Christine", "RS Neo",          3,   30,  50),
    ("Christine", "RS Neo",          4,   27,  18),
    ("Christine", "RS Neo",          5,   28,  19),

    ("Davy",      "Laser",           1,   23,   8),
    ("Davy",      "Laser",           2,   30,  51),
    ("Davy",      "Laser",           3,   30,  50),
    ("Davy",      "Laser",           4,   30,   0),
    ("Davy",      "Laser",           5,   30,   0),
]


# ==============================================================================
# MODULE-FUNCTIES
# ==============================================================================

def load_boat_py_table(extra_csv: str | None = None) -> dict[str, float]:
    """
    Geeft de interne PY-tabel terug.
    Optioneel: laad extra boottypes vanuit een CSV met kolommen 'boottype' en 'py'.
    """
    table = dict(BOAT_PY)
    if extra_csv and Path(extra_csv).exists():
        df = pd.read_csv(extra_csv)
        df.columns = [c.strip().lower() for c in df.columns]
        for _, row in df.iterrows():
            table[str(row["boottype"]).strip()] = float(row["py"])
        print(f"  Extra PY-waarden geladen uit: {extra_csv}")
    return table


def load_race_data(path: str | None = None) -> pd.DataFrame:
    """
    Laad wedstrijddata vanuit CSV of Excel.
    Verwachte kolommen (hoofdletterongevoelig):
        naam, boottype, reeks, minuten, seconden
    Optionele kolom: py  (manuele override van PY-waarde)
    Als 'path' None is, wordt de ingebouwde demo-data gebruikt.
    """
    if path is None:
        df = pd.DataFrame(DEMO_DATA, columns=["naam", "boottype", "reeks", "minuten", "seconden"])
        print("  Demo-data geladen (geen invoerbestand opgegeven).")
        return df

    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Invoerbestand niet gevonden: {path}")

    ext = path.suffix.lower()
    if ext == ".csv":
        df = pd.read_csv(path)
    elif ext in (".xlsx", ".xls"):
        df = pd.read_excel(path)
    else:
        raise ValueError(f"Niet-ondersteund bestandsformaat: {ext}. Gebruik .csv of .xlsx.")

    # Normaliseer kolomnamen
    df.columns = [c.strip().lower() for c in df.columns]
    required = {"naam", "boottype", "reeks", "minuten", "seconden"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Ontbrekende kolommen in invoerbestand: {missing}")

    return df


def calculate_elapsed_seconds(df: pd.DataFrame) -> pd.DataFrame:
    """Voeg kolom 'totaal_seconden' toe op basis van minuten en seconden."""
    df = df.copy()
    df["minuten"] = pd.to_numeric(df["minuten"], errors="coerce")
    df["seconden"] = pd.to_numeric(df["seconden"], errors="coerce")

    ontbrekend = df[df[["minuten", "seconden"]].isna().any(axis=1)]
    if not ontbrekend.empty:
        print(f"  WAARSCHUWING: {len(ontbrekend)} rij(en) met ontbrekende tijden worden overgeslagen:")
        print(ontbrekend[["naam", "reeks"]].to_string(index=False))

    df = df.dropna(subset=["minuten", "seconden"])
    df["totaal_seconden"] = df["minuten"] * 60 + df["seconden"]
    return df


def calculate_corrected_time_py(df: pd.DataFrame, py_table: dict[str, float]) -> pd.DataFrame:
    """
    Bereken gecorrigeerde tijd met de PY-formule:
        gecorrigeerde_tijd = verlopen_tijd_seconden * 1000 / PY

    Als de kolom 'py' aanwezig is in de data, wordt die waarde gebruikt (manuele override).
    Anders wordt de waarde uit py_table gehaald op basis van het boottype.
    """
    df = df.copy()

    def _get_py(row):
        # Manuele override heeft voorrang
        if "py" in df.columns and pd.notna(row.get("py")):
            val = float(row["py"])
            if val <= 0:
                raise ValueError(f"Ongeldige PY-waarde {val} voor {row['naam']}")
            return val
        bt = str(row["boottype"]).strip()
        if bt not in py_table:
            raise KeyError(
                f"Onbekend boottype '{bt}' voor deelnemer '{row['naam']}'. "
                f"Voeg dit type toe aan BOAT_PY of geef een 'py'-kolom op."
            )
        return py_table[bt]

    df["py"] = df.apply(_get_py, axis=1)
    df["gecorrigeerde_tijd"] = df["totaal_seconden"] * 1000 / df["py"]
    return df


def rank_each_race(df: pd.DataFrame) -> pd.DataFrame:
    """
    Voeg een 'rang' kolom toe per reeks op basis van gecorrigeerde tijd (laagste = beste).
    Gelijke tijden krijgen dezelfde rang (dense ranking).
    """
    df = df.copy()
    df["rang"] = df.groupby("reeks")["gecorrigeerde_tijd"].rank(method="min").astype(int)
    return df


def calculate_points(df: pd.DataFrame) -> pd.DataFrame:
    """
    Wijs punten toe: rang = punten (1e = 1 punt, laagste totaal wint).
    Niet-gestarte (DNS) of niet-gefinisht (DNF) deelnemers krijgen aantal_deelnemers + 1 punten
    maar die logica is hier niet nodig tenzij DNS/DNF-rijen aanwezig zijn.
    """
    df = df.copy()
    df["punten"] = df["rang"]
    return df


def drop_worst_result(df: pd.DataFrame) -> pd.DataFrame:
    """
    Schrap het slechtste resultaat per deelnemer (hoogste punten).
    Geeft een nieuw DataFrame terug met de schrapresultaten gemarkeerd in kolom 'geschrapt'.
    """
    df = df.copy()
    df["geschrapt"] = False

    worst_indices = (
        df.groupby("naam")["punten"]
        .idxmax()
        .values
    )
    df.loc[worst_indices, "geschrapt"] = True
    return df


def generate_summary_tables(df: pd.DataFrame, use_drop: bool = False) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Genereer twee tabellen:
    1. detail_tabel: per deelnemer per reeks alle tijden, gecorrigeerde tijd en rang
    2. samenvatting: totaalklassement met optionele schrap

    Geeft (detail_tabel, samenvatting) terug.
    """
    # --- Detailtabel ---
    detail = df[["naam", "boottype", "py", "reeks", "minuten", "seconden",
                 "totaal_seconden", "gecorrigeerde_tijd", "rang", "punten"]].copy()
    detail = detail.rename(columns={
        "naam": "Naam",
        "boottype": "Boottype",
        "py": "PY",
        "reeks": "Reeks",
        "minuten": "Min",
        "seconden": "Sec",
        "totaal_seconden": "Totaal sec",
        "gecorrigeerde_tijd": "Gecorr. tijd",
        "rang": "Rang",
        "punten": "Punten",
    })
    detail = detail.sort_values(["Reeks", "Rang"])

    # --- Samenvatting per deelnemer ---
    reeksen = sorted(df["reeks"].unique())
    rijen = []

    for naam, groep in df.groupby("naam"):
        rij = {"Naam": naam, "Boottype": groep["boottype"].iloc[0]}

        punten_per_reeks = {}
        for reeks in reeksen:
            r = groep[groep["reeks"] == reeks]
            if r.empty:
                punten_per_reeks[reeks] = None
            else:
                punten_per_reeks[reeks] = int(r["punten"].iloc[0])

        for reeks in reeksen:
            rij[f"R{reeks}"] = punten_per_reeks[reeks] if punten_per_reeks[reeks] is not None else "DNS"

        geldige = [p for p in punten_per_reeks.values() if p is not None]
        rij["Som alle"] = sum(geldige)
        rij["Slechtste"] = max(geldige) if geldige else 0

        if use_drop and len(geldige) > 1:
            rij["Totaal punten"] = rij["Som alle"] - rij["Slechtste"]
        else:
            rij["Totaal punten"] = rij["Som alle"]

        rijen.append(rij)

    samenvatting = pd.DataFrame(rijen)
    samenvatting = samenvatting.sort_values("Totaal punten")
    samenvatting.insert(2, "Eindstand", range(1, len(samenvatting) + 1))
    return detail, samenvatting


def export_to_excel(detail: pd.DataFrame, samenvatting: pd.DataFrame,
                    output_path: str = "wedstrijd_resultaten.xlsx") -> str:
    """
    Exporteer detail- en samenvattingstabel naar een opgemaakte Excel-werkmap.
    Geeft het pad naar het aangemaakt bestand terug.
    """
    wb = openpyxl.Workbook()

    _write_samenvatting_sheet(wb, samenvatting)
    _write_detail_sheet(wb, detail)
    _write_rangschikking_sheet(wb, samenvatting)

    # Verwijder default leeg blad
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    return output_path


# ---- Hulpfuncties voor Excel-opmaak ----

def _header_stijl(ws, row: int, fill_hex: str = "1F4E79"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _border_all(ws):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3


def _write_samenvatting_sheet(wb, samenvatting: pd.DataFrame):
    ws = wb.create_sheet("Klassement")
    ws.title = "Klassement"

    ws.append(["TOTAALKLASSEMENT"])
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.append([])

    headers = list(samenvatting.columns)
    ws.append(headers)
    _header_stijl(ws, 3)

    for _, row in samenvatting.iterrows():
        ws.append(list(row))

    # Kleuren top-3
    kleuren = ["FFD700", "C0C0C0", "CD7F32"]  # goud, zilver, brons
    for i, kleur in enumerate(kleuren):
        data_row = 4 + i
        if data_row <= ws.max_row:
            fill = PatternFill("solid", fgColor=kleur)
            for cell in ws[data_row]:
                cell.fill = fill

    _border_all(ws)
    _auto_width(ws)


def _write_detail_sheet(wb, detail: pd.DataFrame):
    ws = wb.create_sheet("Detail per reeks")

    ws.append(["DETAIL PER REEKS"])
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.append([])

    reeksen = sorted(detail["Reeks"].unique())
    for reeks in reeksen:
        ws.append([f"Reeks {reeks}"])
        hdr_row = ws.max_row
        ws[f"A{hdr_row}"].font = Font(bold=True, size=12, color="2E75B6")

        subset = detail[detail["Reeks"] == reeks].drop(columns=["Reeks"])
        headers = list(subset.columns)
        ws.append(headers)
        _header_stijl(ws, ws.max_row, fill_hex="2E75B6")

        for _, row in subset.iterrows():
            vals = []
            for v in row:
                if isinstance(v, float):
                    vals.append(round(v, 2))
                else:
                    vals.append(v)
            ws.append(vals)

        ws.append([])  # lege rij tussen reeksen

    _border_all(ws)
    _auto_width(ws)


def _write_rangschikking_sheet(wb, samenvatting: pd.DataFrame):
    ws = wb.create_sheet("Posities per reeks")

    ws.append(["POSITIES PER REEKS"])
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.append([])

    reeks_cols = [c for c in samenvatting.columns if c.startswith("R") and c[1:].isdigit()]
    headers = ["Eindstand", "Naam", "Boottype"] + reeks_cols + ["Totaal punten"]
    ws.append(headers)
    _header_stijl(ws, 3)

    for _, row in samenvatting.iterrows():
        ws.append([row.get("Eindstand"), row.get("Naam"), row.get("Boottype")]
                  + [row.get(r) for r in reeks_cols]
                  + [row.get("Totaal punten")])

    _border_all(ws)
    _auto_width(ws)


# ==============================================================================
# HOOFD-PIJPLIJN
# ==============================================================================

def bereken_wedstrijd(invoer_pad: str | None, schrap: bool = True,
                      uitvoer_pad: str = "wedstrijd_resultaten.xlsx",
                      extra_py_csv: str | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Volledige berekeningspijplijn.
    Geeft (detail, samenvatting) terug en schrijft een Excel-bestand.
    """
    print("\n=== WedstrijdCalculator - PY Scoring ===\n")

    py_table = load_boat_py_table(extra_py_csv)
    print(f"  {len(py_table)} boottypes geladen in PY-tabel.")

    df = load_race_data(invoer_pad)
    print(f"  {len(df)} rijen geladen ({df['naam'].nunique()} deelnemers, "
          f"{df['reeks'].nunique()} reeksen).")

    df = calculate_elapsed_seconds(df)
    df = calculate_corrected_time_py(df, py_table)
    df = rank_each_race(df)
    df = calculate_points(df)

    if schrap:
        df = drop_worst_result(df)
        print("  Schrapresultaat: slechtste reeks per deelnemer niet meegeteld.")

    detail, samenvatting = generate_summary_tables(df, use_drop=schrap)

    pad = export_to_excel(detail, samenvatting, uitvoer_pad)
    print(f"\n  Excel-export opgeslagen: {pad}")

    return detail, samenvatting


def druk_klassement(samenvatting: pd.DataFrame):
    """Druk het klassement overzichtelijk af in de terminal."""
    print("\n" + "=" * 60)
    print("  EINDKLASSEMENT")
    print("=" * 60)
    print(samenvatting.to_string(index=False))
    print("=" * 60)


# ==============================================================================
# CLI
# ==============================================================================

def interactief_menu():
    print("\n╔══════════════════════════════════════════╗")
    print("║  WedstrijdCalculator - PY Handicap Scoring ║")
    print("╚══════════════════════════════════════════╝\n")
    print("Opties:")
    print("  1. Verwerk invoerbestand (CSV of Excel)")
    print("  2. Start met demo-data")
    print("  3. Toon beschikbare boottypes en PY-waarden")
    print("  4. Afsluiten")
    keuze = input("\nKeuze (1-4): ").strip()

    if keuze == "1":
        pad = input("Pad naar invoerbestand (CSV/XLSX): ").strip()
        schrap_str = input("Schrapresultaat toepassen? (j/n, standaard j): ").strip().lower()
        schrap = schrap_str != "n"
        uitvoer = input("Uitvoerbestand (Enter = wedstrijd_resultaten.xlsx): ").strip()
        if not uitvoer:
            uitvoer = "wedstrijd_resultaten.xlsx"
        detail, samenvatting = bereken_wedstrijd(pad, schrap=schrap, uitvoer_pad=uitvoer)
        druk_klassement(samenvatting)

    elif keuze == "2":
        schrap_str = input("Schrapresultaat toepassen? (j/n, standaard j): ").strip().lower()
        schrap = schrap_str != "n"
        detail, samenvatting = bereken_wedstrijd(None, schrap=schrap)
        druk_klassement(samenvatting)

    elif keuze == "3":
        py_table = load_boat_py_table()
        print(f"\n{'Boottype':<25} {'PY':>6}")
        print("-" * 33)
        for bt, py in sorted(py_table.items(), key=lambda x: x[1]):
            print(f"{bt:<25} {py:>6.0f}")
        print()

    elif keuze == "4":
        print("Tot ziens!")
        sys.exit(0)
    else:
        print("Ongeldige keuze.")


def main():
    parser = argparse.ArgumentParser(
        description="WedstrijdCalculator - Portsmouth Yardstick scoring tool"
    )
    parser.add_argument("--invoer", "-i", help="Pad naar CSV- of Excel-invoerbestand")
    parser.add_argument("--uitvoer", "-o", default="wedstrijd_resultaten.xlsx",
                        help="Pad voor Excel-uitvoer (standaard: wedstrijd_resultaten.xlsx)")
    parser.add_argument("--schrap", action="store_true", default=True,
                        help="Schrap slechtste reeks per deelnemer (standaard: aan)")
    parser.add_argument("--geen-schrap", dest="schrap", action="store_false",
                        help="Geen schrapresultaat toepassen")
    parser.add_argument("--extra-py", help="CSV-bestand met extra boottype/PY-waarden")
    parser.add_argument("--demo", action="store_true",
                        help="Voer uit met ingebouwde demo-data")
    args = parser.parse_args()

    if args.demo:
        detail, samenvatting = bereken_wedstrijd(
            None, schrap=args.schrap, uitvoer_pad=args.uitvoer, extra_py_csv=args.extra_py
        )
        druk_klassement(samenvatting)
    elif args.invoer:
        detail, samenvatting = bereken_wedstrijd(
            args.invoer, schrap=args.schrap, uitvoer_pad=args.uitvoer, extra_py_csv=args.extra_py
        )
        druk_klassement(samenvatting)
    else:
        # Geen argumenten: interactief menu
        interactief_menu()


if __name__ == "__main__":
    main()
